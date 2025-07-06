import openpyxl

var1 = openpyxl.load_workbook("inventory.xlsx")
product_list = var1["Sheet1"]

product_per_suplier = {}

# print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value

    if supplier_name in product_per_suplier:
        current_number_product = product_per_suplier.get(supplier_name)
        product_per_suplier[supplier_name] = current_number_product + 1
    else:
        product_per_suplier[supplier_name] = 1

print(product_per_suplier)
